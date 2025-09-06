"""
Analytics utilities for the vehicle service system.
Provides data aggregation and analysis functions for dashboards and reports.
"""

from django.db.models import Count, Q, Avg
from django.utils import timezone
from datetime import datetime, timedelta
from collections import defaultdict
import json


class ServiceAnalytics:
    """Analytics class for service data analysis."""
    
    @staticmethod
    def get_monthly_service_counts(months=12):
        """Get service counts for the last N months."""
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=months * 30)
        
        from .models import Service
        
        # Get services grouped by month
        services = Service.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).extra(
            select={'month': "strftime('%%Y-%%m', created_at)"}
        ).values('month').annotate(count=Count('id')).order_by('month')
        
        # Create a complete month range
        monthly_data = {}
        current = start_date.replace(day=1)
        while current <= end_date:
            month_key = current.strftime('%Y-%m')
            monthly_data[month_key] = 0
            current = (current + timedelta(days=32)).replace(day=1)
        
        # Fill in actual data
        for service in services:
            monthly_data[service['month']] = service['count']
        
        return monthly_data
    
    @staticmethod
    def get_service_status_distribution():
        """Get distribution of service statuses."""
        from .models import Service
        
        status_counts = Service.objects.values('status').annotate(count=Count('id'))
        return {item['status']: item['count'] for item in status_counts}
    
    @staticmethod
    def get_service_type_distribution():
        """Get distribution of service types."""
        from .models import Service
        
        type_counts = Service.objects.values('service_type').annotate(count=Count('id'))
        return {item['service_type']: item['count'] for item in type_counts}
    
    @staticmethod
    def get_mechanic_performance():
        """Get performance metrics for mechanics."""
        from .models import Service, Profile
        
        mechanics = Profile.objects.filter(role=Profile.ROLE_MECHANIC)
        performance_data = []
        
        for mechanic in mechanics:
            assigned_services = Service.objects.filter(assigned_mechanic=mechanic.user)
            completed_services = assigned_services.filter(status=Service.STATUS_COMPLETED)
            pending_services = assigned_services.filter(status=Service.STATUS_PENDING)
            in_progress_services = assigned_services.filter(status=Service.STATUS_IN_PROGRESS)
            
            performance_data.append({
                'mechanic_name': mechanic.user.get_full_name() or mechanic.user.username,
                'total_assigned': assigned_services.count(),
                'completed': completed_services.count(),
                'pending': pending_services.count(),
                'in_progress': in_progress_services.count(),
                'completion_rate': (completed_services.count() / assigned_services.count() * 100) if assigned_services.count() > 0 else 0
            })
        
        return performance_data
    
    @staticmethod
    def get_customer_service_history(user):
        """Get service history for a specific customer."""
        from .models import Service
        
        # Get services by year
        services = Service.objects.filter(customer=user).extra(
            select={'year': "strftime('%%Y', created_at)"}
        ).values('year', 'status').annotate(count=Count('id'))
        
        yearly_data = defaultdict(lambda: {'completed': 0, 'pending': 0, 'in_progress': 0})
        
        for service in services:
            year = service['year']
            status = service['status']
            count = service['count']
            
            if status == Service.STATUS_COMPLETED:
                yearly_data[year]['completed'] = count
            elif status == Service.STATUS_PENDING:
                yearly_data[year]['pending'] = count
            elif status == Service.STATUS_IN_PROGRESS:
                yearly_data[year]['in_progress'] = count
        
        return dict(yearly_data)
    
    @staticmethod
    def get_manager_insights():
        """Get comprehensive insights for managers."""
        from .models import Service, Profile
        
        # Basic counts
        total_services = Service.objects.count()
        pending_services = Service.objects.filter(status=Service.STATUS_PENDING).count()
        completed_services = Service.objects.filter(status=Service.STATUS_COMPLETED).count()
        in_progress_services = Service.objects.filter(status=Service.STATUS_IN_PROGRESS).count()
        
        # This month's data
        now = timezone.now()
        this_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        this_month_services = Service.objects.filter(created_at__gte=this_month_start).count()
        
        # Top service types
        top_service_types = Service.objects.values('service_type').annotate(
            count=Count('id')
        ).order_by('-count')[:3]
        
        # Format service type names for display
        for service_type in top_service_types:
            service_type['display_name'] = service_type['service_type'].replace('_', ' ').title()
        
        # Busiest mechanics
        busiest_mechanics = Service.objects.filter(
            assigned_mechanic__isnull=False
        ).values(
            'assigned_mechanic__first_name',
            'assigned_mechanic__last_name',
            'assigned_mechanic__username'
        ).annotate(
            total_jobs=Count('id')
        ).order_by('-total_jobs')[:3]
        
        return {
            'total_services': total_services,
            'pending_services': pending_services,
            'completed_services': completed_services,
            'in_progress_services': in_progress_services,
            'this_month_services': this_month_services,
            'completion_rate': (completed_services / total_services * 100) if total_services > 0 else 0,
            'top_service_types': list(top_service_types),
            'busiest_mechanics': list(busiest_mechanics),
            'monthly_data': ServiceAnalytics.get_monthly_service_counts(),
            'status_distribution': ServiceAnalytics.get_service_status_distribution(),
            'service_type_distribution': ServiceAnalytics.get_service_type_distribution()
        }
    
    @staticmethod
    def get_mechanic_insights(mechanic_user):
        """Get insights for a specific mechanic."""
        from .models import Service
        
        assigned_services = Service.objects.filter(assigned_mechanic=mechanic_user)
        completed_services = assigned_services.filter(status=Service.STATUS_COMPLETED)
        pending_services = assigned_services.filter(status=Service.STATUS_PENDING)
        in_progress_services = assigned_services.filter(status=Service.STATUS_IN_PROGRESS)
        
        # This month's performance
        now = timezone.now()
        this_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        this_month_completed = completed_services.filter(created_at__gte=this_month_start).count()
        
        return {
            'total_assigned': assigned_services.count(),
            'completed': completed_services.count(),
            'pending': pending_services.count(),
            'in_progress': in_progress_services.count(),
            'this_month_completed': this_month_completed,
            'completion_rate': (completed_services.count() / assigned_services.count() * 100) if assigned_services.count() > 0 else 0
        }
    
    @staticmethod
    def get_customer_insights(customer_user):
        """Get insights for a specific customer."""
        from .models import Service
        
        customer_services = Service.objects.filter(customer=customer_user)
        completed_services = customer_services.filter(status=Service.STATUS_COMPLETED)
        pending_services = customer_services.filter(status=Service.STATUS_PENDING)
        in_progress_services = customer_services.filter(status=Service.STATUS_IN_PROGRESS)
        
        # This year's data
        now = timezone.now()
        this_year_start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        this_year_services = customer_services.filter(created_at__gte=this_year_start).count()
        
        return {
            'total_services': customer_services.count(),
            'completed': completed_services.count(),
            'pending': pending_services.count(),
            'in_progress': in_progress_services.count(),
            'this_year_services': this_year_services,
            'service_history': ServiceAnalytics.get_customer_service_history(customer_user)
        }


class ReportGenerator:
    """Generate various types of reports."""
    
    @staticmethod
    def generate_csv_report(services_queryset, filename="services_report.csv"):
        """Generate CSV report from services queryset."""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Customer', 'Service Type', 'Status', 'Assigned Mechanic', 
            'Created Date'
        ])
        
        for service in services_queryset:
            writer.writerow([
                service.id,
                service.customer.get_full_name() or service.customer.username,
                service.get_service_type_display(),
                service.get_status_display(),
                service.assigned_mechanic.get_full_name() if service.assigned_mechanic else 'Unassigned',
                service.created_at.strftime('%Y-%m-%d %H:%M')
            ])
        
        return response
    
    @staticmethod
    def generate_excel_report(services_queryset, filename="services_report.xlsx"):
        """Generate Excel report from services queryset."""
        try:
            import openpyxl
            from django.http import HttpResponse
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Services Report"
            
            # Headers
            headers = [
                'ID', 'Customer', 'Service Type', 'Status', 'Assigned Mechanic', 
                'Created Date'
            ]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Data
            for row, service in enumerate(services_queryset, 2):
                ws.cell(row=row, column=1, value=service.id)
                ws.cell(row=row, column=2, value=service.customer.get_full_name() or service.customer.username)
                ws.cell(row=row, column=3, value=service.get_service_type_display())
                ws.cell(row=row, column=4, value=service.get_status_display())
                ws.cell(row=row, column=5, value=service.assigned_mechanic.get_full_name() if service.assigned_mechanic else 'Unassigned')
                ws.cell(row=row, column=6, value=service.created_at.strftime('%Y-%m-%d %H:%M'))
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
            
        except ImportError:
            # Fallback to CSV if openpyxl is not available
            return ReportGenerator.generate_csv_report(services_queryset, filename.replace('.xlsx', '.csv'))
